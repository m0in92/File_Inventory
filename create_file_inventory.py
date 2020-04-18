import openpyxl as xls
from pathlib import Path
import xlsxwriter

workbook = xlsxwriter.Workbook('File_Inventory.xlsx')
worksheet = workbook.add_worksheet()
workbook.close()

wb = xls.load_workbook("File_Inventory.xlsx")
sheet = wb['Sheet1']
filename_heading_cell = sheet.cell(1,1)
filetype_heading_cell = sheet.cell(1,2)
filename_heading_cell.value = 'File Name'
filetype_heading_cell.value = 'File Type'

def filename():
    path=Path()
    filenames=[]
    filetypes=[]
    for files in path.glob('*'):
        filenames.append(str(files).split('.')[0])
    return filenames

def filetypes():
    path=Path()
    filenames=[]
    filetypes=[]
    for files in path.glob('*'):
        if '.' in str(files):
            filetypes.append(str(files).split('.')[1])
        else:
            filetypes.append('directory')
    return filetypes

filenames = filename()
filetypes = filetypes()
for items in range(2,len(filenames)):
    filenames_row_value=sheet.cell(items,1)
    filetypes_row_value = sheet.cell(items, 2)
    filenames_row_value.value=filenames[items]
    filetypes_row_value.value = filetypes[items]

wb.save('File_Inventory.xlsx')